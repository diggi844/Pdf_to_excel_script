#!/usr/bin/env python3
"""
giornale_to_excel.py
=====================
Converts an Italian accounting ledger ("Giornale") PDF into a structured Excel file.

Target columns (in order):
    Data | Comp. | Descrizione Movimento | Documento | N. del e Prot. | Partita | DARE | AVERE | SALDO

Design notes
------------
- 100% local / offline. Only pdfplumber (extraction) + openpyxl (writing). No network calls,
  safe for confidential client data.
- Uses a two-strategy extraction approach per page:
    1) Table extraction via pdfplumber's built-in table detector (works when the PDF has
       real ruling lines / consistent whitespace-based columns).
    2) Fallback: word-position (coordinate) clustering into rows/columns when table
       detection fails or returns too few columns.
- Content-driven row classification: rather than trusting the table grid blindly, each
  extracted row is validated against expected patterns (date regex, numeric regex for
  DARE/AVERE/SALDO) before being accepted as a transaction row.
- Skips the "page-end / carry-forward totals" row: on many of these ledgers, at the
  bottom of each page there's a summary row where BOTH DARE and AVERE are populated
  together (a subtotal / "riporto"), with no real Descrizione/Documento content.
  This is distinguished from genuine transaction rows (which normally populate only
  one of DARE/AVERE) and is filtered out automatically. See `is_page_total_row()`.
- Blank DARE/AVERE default to 0 (as int/float 0, not empty string) so Excel can sum them.
- Multi-page: processes every page in the PDF, not just page 1, but page 1 is treated
  specially since it usually also carries the account header info (kept for reference/logging).

Usage
-----
    python giornale_to_excel.py input.pdf output.xlsx

If output.xlsx is omitted, it defaults to the same name as input.pdf with .xlsx extension.
"""

import sys
import re
import argparse
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------------------
# Column definitions
# --------------------------------------------------------------------------------------
COLUMNS = [
    "Data",
    "Comp.",
    "Descrizione Movimento",
    "Documento",
    "N. del e Prot.",
    "Partita",
    "DARE",
    "AVERE",
    "SALDO",
]

# --------------------------------------------------------------------------------------
# Regex helpers
# --------------------------------------------------------------------------------------
# Italian date formats: 01/02/2024, 01-02-2024, 01.02.2024, 1/2/24
DATE_RE = re.compile(r"\b(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4})\b")

# Italian-formatted numbers: 1.234,56  or  1234,56  or  -1.234,56  or with trailing sign
NUM_RE = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}-?")

# A "protocol / N del" style token: numbers combined with dates, e.g. "123 del 01/02/2024"
NRIF_RE = re.compile(r"\d+\s*(?:/\s*\d+)?\s*(?:del)?\s*\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4}", re.IGNORECASE)


def parse_it_number(token):
    """Convert an Italian-formatted number string ('1.234,56' or '1234,56-') to float.
    Returns None if it doesn't look like a number."""
    if token is None:
        return None
    token = token.strip()
    if token == "" or token in {"-", "--"}:
        return None
    negative = token.endswith("-")
    token = token.rstrip("-")
    token = token.replace(".", "").replace(",", ".")
    try:
        val = float(token)
        return -val if negative else val
    except ValueError:
        return None


def looks_like_date(token):
    return bool(DATE_RE.search(token or ""))


def is_page_total_row(row_dict):
    """
    Detects the page-end carry-forward / subtotal row that should be SKIPPED.

    Heuristic: on these ledgers, a genuine transaction populates DARE *or* AVERE
    (rarely both), and always carries a Data + Descrizione. The page-end totals
    row instead has BOTH DARE and AVERE populated together, while Data /
    Descrizione / Documento are empty or contain only generic keywords like
    "riporto", "totale", "saldo a riportare", etc.
    """
    dare = row_dict.get("DARE")
    avere = row_dict.get("AVERE")
    data = (row_dict.get("Data") or "").strip()
    descr = (row_dict.get("Descrizione Movimento") or "").strip().lower()

    both_populated = (dare not in (None, 0, 0.0)) and (avere not in (None, 0, 0.0))

    keyword_hit = any(
        kw in descr
        for kw in ("riporto", "riportare", "totale", "totali", "saldo prec", "a riportare")
    )

    if both_populated and (data == "" or keyword_hit):
        return True

    # Also catch rows where Data is empty AND descrizione is empty AND both DARE/AVERE set
    if both_populated and data == "" and descr == "":
        return True

    return False


# --------------------------------------------------------------------------------------
# Strategy 1: table extraction
# --------------------------------------------------------------------------------------
def extract_via_tables(page):
    """Try pdfplumber's table detection. Returns a list of row dicts, or None if it
    doesn't produce usable results."""
    table_settings_options = [
        {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
        {"vertical_strategy": "text", "horizontal_strategy": "text"},
        {"vertical_strategy": "lines", "horizontal_strategy": "text"},
    ]

    for settings in table_settings_options:
        try:
            tables = page.extract_tables(settings)
        except Exception:
            continue
        if not tables:
            continue

        for table in tables:
            rows = []
            for raw_row in table:
                if not raw_row or all((c is None or str(c).strip() == "") for c in raw_row):
                    continue
                row_dict = map_raw_row_to_columns(raw_row)
                if row_dict:
                    rows.append(row_dict)
            if len(rows) >= 1:
                return rows
    return None


def map_raw_row_to_columns(raw_row):
    """Map a raw list-of-cells row (variable length/order) onto our target COLUMNS
    using content-driven matching rather than assuming fixed positions."""
    cells = [(c or "").strip() for c in raw_row]
    # drop fully empty rows
    if all(c == "" for c in cells):
        return None

    row = {col: "" for col in COLUMNS}

    # Identify numeric cells (candidates for DARE / AVERE / SALDO) from the right side
    numeric_idxs = [i for i, c in enumerate(cells) if NUM_RE.fullmatch(c.replace(" ", ""))]

    # Identify date cell (Data) - usually leftmost date match
    date_idx = next((i for i, c in enumerate(cells) if looks_like_date(c)), None)

    if date_idx is not None:
        row["Data"] = DATE_RE.search(cells[date_idx]).group(1)

    # Assign trailing numeric cells to DARE/AVERE/SALDO in order of appearance (left->right)
    # Typical layout: ... DARE | AVERE | SALDO  (last up to 3 numeric cells)
    trailing_numeric = numeric_idxs[-3:] if len(numeric_idxs) >= 1 else []
    numeric_values = [parse_it_number(cells[i]) for i in trailing_numeric]

    if len(numeric_values) == 3:
        row["DARE"], row["AVERE"], row["SALDO"] = numeric_values
    elif len(numeric_values) == 2:
        # Ambiguous: could be (DARE, SALDO) or (AVERE, SALDO) or (DARE, AVERE)
        row["DARE"], row["AVERE"] = numeric_values[0], None
        row["SALDO"] = numeric_values[1]
    elif len(numeric_values) == 1:
        row["SALDO"] = numeric_values[0]

    # Everything else (excluding date_idx and numeric idxs) becomes descriptive text
    excluded = set(numeric_idxs) | ({date_idx} if date_idx is not None else set())
    remaining_texts = [cells[i] for i in range(len(cells)) if i not in excluded and cells[i]]

    # Try to find "N. del e Prot." style token (number + date combo) among remaining texts
    nrif_idx_in_remaining = None
    for i, txt in enumerate(remaining_texts):
        if NRIF_RE.search(txt):
            nrif_idx_in_remaining = i
            break
    if nrif_idx_in_remaining is not None:
        row["N. del e Prot."] = remaining_texts.pop(nrif_idx_in_remaining)

    # Heuristic assignment of remaining text fields, left to right:
    # Comp. (short code) -> Descrizione Movimento (longest text) -> Documento -> Partita
    if remaining_texts:
        # The longest string is very likely the Descrizione Movimento
        descr_idx = max(range(len(remaining_texts)), key=lambda i: len(remaining_texts[i]))
        row["Descrizione Movimento"] = remaining_texts.pop(descr_idx)

    if remaining_texts:
        # Short (<=6 chars) alphanumeric token -> Comp.
        comp_idx = next((i for i, t in enumerate(remaining_texts) if len(t) <= 6), None)
        if comp_idx is not None:
            row["Comp."] = remaining_texts.pop(comp_idx)

    if remaining_texts:
        row["Documento"] = remaining_texts.pop(0)

    if remaining_texts:
        row["Partita"] = remaining_texts.pop(0)

    # Default blank DARE/AVERE to 0
    row["DARE"] = row["DARE"] if row["DARE"] not in (None, "") else 0
    row["AVERE"] = row["AVERE"] if row["AVERE"] not in (None, "") else 0

    return row


# --------------------------------------------------------------------------------------
# Strategy 2: coordinate/word clustering fallback
# --------------------------------------------------------------------------------------
def extract_via_words(page):
    """Fallback strategy: cluster words into lines by vertical position (`top`), then
    split each line into columns by horizontal gaps. Useful when no ruled table exists."""
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    if not words:
        return []

    # Group words into lines by rounding 'top' to nearest few pixels
    lines = {}
    for w in words:
        key = round(w["top"] / 3) * 3  # tolerance band
        lines.setdefault(key, []).append(w)

    rows = []
    for key in sorted(lines.keys()):
        line_words = sorted(lines[key], key=lambda w: w["x0"])
        # Merge words into "cells" based on horizontal gap threshold
        cells = []
        current_cell = [line_words[0]]
        for prev, curr in zip(line_words, line_words[1:]):
            gap = curr["x0"] - prev["x1"]
            if gap > 8:  # gap threshold indicating new column
                cells.append(" ".join(w["text"] for w in current_cell))
                current_cell = [curr]
            else:
                current_cell.append(curr)
        cells.append(" ".join(w["text"] for w in current_cell))

        row_dict = map_raw_row_to_columns(cells)
        if row_dict:
            rows.append(row_dict)

    return rows


# --------------------------------------------------------------------------------------
# Main per-page extraction orchestration
# --------------------------------------------------------------------------------------
def extract_page(page, page_num):
    rows = extract_via_tables(page)
    if not rows:
        rows = extract_via_words(page)

    valid_rows = []
    skipped_totals = 0
    for row in rows:
        # Skip separator / empty rows: must have at least a date OR a description
        if not row.get("Data") and not row.get("Descrizione Movimento"):
            continue

        if is_page_total_row(row):
            skipped_totals += 1
            continue

        valid_rows.append(row)

    if skipped_totals:
        print(f"  [page {page_num}] skipped {skipped_totals} page-total/carry-forward row(s)")

    return valid_rows


def process_pdf(pdf_path):
    all_rows = []
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        print(f"Opened '{pdf_path}' — {total_pages} page(s) detected.")

        for i, page in enumerate(pdf.pages, start=1):
            print(f"Processing page {i}/{total_pages}...")
            page_rows = extract_page(page, i)
            print(f"  -> extracted {len(page_rows)} transaction row(s)")
            all_rows.extend(page_rows)

    return all_rows


# --------------------------------------------------------------------------------------
# Excel writing
# --------------------------------------------------------------------------------------
def write_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Giornale"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col_name in enumerate(COLUMNS, start=1):
            value = row.get(col_name, "")
            if col_name in ("DARE", "AVERE", "SALDO"):
                value = value if value not in (None, "") else 0
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Auto-fit column widths (approximate)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        max_len = len(col_name)
        for row in rows:
            val = row.get(col_name, "")
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\nSaved {len(rows)} row(s) to '{output_path}'")


# --------------------------------------------------------------------------------------
# CLI entry point
# --------------------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Convert Italian ledger (Giornale) PDF to Excel.")
    parser.add_argument("input_pdf", help="Path to the input PDF file")
    parser.add_argument("output_xlsx", nargs="?", help="Path to the output Excel file (optional)")
    args = parser.parse_args()

    input_path = Path(args.input_pdf)
    if not input_path.exists():
        print(f"ERROR: file not found: {input_path}")
        sys.exit(1)

    output_path = Path(args.output_xlsx) if args.output_xlsx else input_path.with_suffix(".xlsx")

    rows = process_pdf(str(input_path))
    if not rows:
        print("WARNING: No transaction rows were extracted. The PDF layout may need custom tuning.")
    write_excel(rows, str(output_path))


if __name__ == "__main__":
    main()
