#!/usr/bin/env python3
"""
giornale_to_excel.py
=====================
Converts an Italian accounting ledger ("Giornale") PDF into a structured Excel file.

REAL structure confirmed from extract_tables() output on the actual PDF:
  extract_tables() splits the HEADER row cleanly into 8 columns:
      ['Data Comp.', 'Registr. N. del', 'Descrizione Movimento',
       'Documento N. del e Prot.', 'Partita', 'DARE', 'AVERE', 'SALDO']
  ...but DATA rows collapse into a single blob in column 0, with every other
  column empty. Each physical line inside that blob is one transaction, e.g.:

      29/05/2025 ***** 29/05/2025 AUMENTO DI CAPITALE COME DA ASSEMBLEA 500.000,00 500.000,00
      09/09/2025 ***** 09/09/2025 sottoscrizione aumento di capitale [PERSON_1] 250.000,00 250.000,00
      09/09/2025 ***** 09/09/2025 sottoscrizione aumento di capitale [PERSON_2] 250.000,00
      500.000,00 500.000,00                      <- orphan line, no date: page-total row, SKIP

  So this script does NOT try to trust extract_tables()'s column splitting for data
  rows. Instead it takes the raw text blob and regex-parses each line itself.

Target columns (in order) — EXACTLY matching the real header row extracted from the
PDF (row[0] of extract_tables() output), not a custom split:
    Data Comp. | Registr. N° del | Descrizione Movimento | Documento N° del e Prot.
    | Partita | DARE | AVERE | SALDO

Design notes
------------
- 100% local / offline. Only pdfplumber (extraction) + openpyxl (writing). No network
  calls, safe for confidential client data.
- Extraction: pdfplumber's extract_tables() (lines/lines, text/text, lines/text tried
  in order) to get the header + locate the data blob; then per-line TOKEN-BASED
  parsing of the blob text itself (see parse_transaction_line()) — not one rigid
  regex, since Documento/Partita are optional per row.
- Field rules per transaction line:
    * Data:                  first date token, kept as-is.
    * Registr. N° del:       number (dots stripped, '8.269' -> '8269') or '*****'
                              placeholder + a date -> "<num> <date>".
    * Descrizione Movimento: free text in between, kept as one joined string.
    * Documento N° del e Prot. (optional): [word][date][number, dots stripped,
                              '2.66' -> '266'] -> one joined string, blank if absent.
    * Partita (optional):    single token 'alnum/yyyy', blank if absent.
    * DARE / AVERE / SALDO:  trailing Italian money tokens ('.'=thousands,
                              ','=decimal, '57.869,10' -> 57869.10).
- DARE vs AVERE assignment: since the single "Importo" figure has no column
  information left once it's inside one text blob, this script infers DARE vs AVERE
  by comparing the transaction's SALDO to the running SALDO carried from the previous
  transaction: if SALDO increases -> DARE, if it decreases -> AVERE.
  ASSUMPTION — please confirm this matches your ledger's convention; it's the most
  common one for Italian "mastrino" client/partner ledgers but can be inverted for
  some account types. See `assign_dare_avere()` — flip the two branches if wrong.
- Missing SALDO recovery: if a transaction line only has 1 trailing amount (SALDO
  missing, likely wrapped onto the next physical line), the script borrows the FIRST
  number of the following orphan line as that transaction's SALDO, then discards the
  rest of the orphan line. Flagged in the console output for manual spot-check.
- Skips a row when BOTH DARE and AVERE are populated together, or when DARE, AVERE,
  AND SALDO are all populated together — the page-end carry-forward/totals row, per
  your explicit instruction. See `is_page_total_row()`.
- Blank fields (Documento, Partita, missing amounts) are left blank/0 rather than
  guessed.
- Processes ALL pages of the PDF by default (use --pages N to limit to first N pages).

Usage
-----
    python giornale_to_excel.py input.pdf output.xlsx      (processes ALL pages)
    python giornale_to_excel.py input.pdf output.xlsx --pages N   (optional: limit to first N pages)

If output.xlsx is omitted, it defaults to the same name as input.pdf with .xlsx extension.
"""

import sys
import re
import argparse
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------------------
# Column definitions (matches the REAL header found via extract_tables())
# --------------------------------------------------------------------------------------
COLUMNS = [
    "Page",
    "Code",
    "Data Comp.",
    "Registr. N° del",
    "Descrizione Movimento",
    "Documento N° del e Prot.",
    "Partita",
    "DARE",
    "AVERE",
    "SALDO",
    "TOTAL_DARE",
    "TOTAL_AVERE",
    "TOTAL_SALDO",
    "Status",
]

# Computed/derived columns — not extracted as row-level zone data, so excluded from
# the extraction field set below. ("Code" is extracted per PAGE, not per row.)
COMPUTED_COLUMNS = ("Page", "Code", "Status", "TOTAL_DARE", "TOTAL_AVERE", "TOTAL_SALDO")

# The columns we actually extract data into (excludes computed columns).
DATA_FIELD_COLUMNS = [c for c in COLUMNS if c not in COMPUTED_COLUMNS]

# Expected header TEXT (word-by-word, in order) per column, used to locate each
# column's header words on the page and derive its x-coordinate zone. Based on the
# real header confirmed via extract_tables(): ['Data Comp.', 'Registr. N° del',
# 'Descrizione Movimento', 'Documento N° del e Prot.', 'Partita', 'DARE', 'AVERE', 'SALDO']
COLUMN_HEADER_WORDS = {
    "Data Comp.": ["Data", "Comp."],
    "Registr. N° del": ["Registr.", "N°", "del"],
    "Descrizione Movimento": ["Descrizione", "Movimento"],
    "Documento N° del e Prot.": ["Documento", "N°", "del", "e", "Prot."],
    "Partita": ["Partita"],
    "DARE": ["DARE"],
    "AVERE": ["AVERE"],
    "SALDO": ["SALDO"],
}

# --------------------------------------------------------------------------------------
# Regex helpers
# --------------------------------------------------------------------------------------
# Header detection keywords (loose match on the actual header wording)
HEADER_KEYWORDS = ["data", "comp", "registr", "descrizione", "movimento",
                   "documento", "prot", "partita", "dare", "avere", "saldo"]

# --- Token-level patterns for each field (see parse_transaction_line() for how
#     these are consumed left-to-right / right-to-left along a line) -------------
DATE_TOKEN_RE = re.compile(r'^\d{1,2}/\d{1,2}/\d{2,4}$')
REGNUM_TOKEN_RE = re.compile(r'^(?:\d[\d.]*|\*+)$')          # e.g. "8.269" or "*****"
DOCNUM_TOKEN_RE = re.compile(r'^\d[\d.]*$')                  # e.g. "2.66" -> "266" (dots stripped, no decimal comma)
PARTITA_TOKEN_RE = re.compile(r'^[A-Za-z0-9-]+(?:/[A-Za-z0-9-]+)+$')  # e.g. "AB123/2025", "26/2025/FE/2025", "4/01-24350/2025"
MONEY_TOKEN_RE = re.compile(r'^-?\d{1,3}(?:\.\d{3})*,\d{2}-?$')  # e.g. "57.869,10" -> 57869.10
MONEY_TOKEN_RE_FINDALL = re.compile(r'-?\d{1,3}(?:\.\d{3})*,\d{2}-?')  # non-anchored, for scanning full lines

# Cap on how many tokens Documento's first field ("word") can span, e.g. "FPR 29/25"
# is 2 tokens. ASSUMPTION based on observed samples — raise if real data needs more.
DOCUMENTO_WORD_MAX_TOKENS = 2

# Sentinel inserted into the tagged-line sequence at every detected header row, used
# to reset the running SALDO (see parse_transaction_lines()) since a repeated header
# likely marks the start of a new account/client/supplier section in the ledger.
HEADER_RESET_SENTINEL = "\x00__HEADER_RESET__\x00"


def parse_it_number(token):
    """Convert an Italian-formatted MONEY string ('1.234,56' or '1234,56-') to float.
    Used for DARE / AVERE / SALDO (dot = thousands separator, comma = decimal)."""
    if token is None:
        return None
    token = token.strip()
    if token == "" or token in {"-", "--"}:
        return None
    negative = token.endswith("-")
    token = token.rstrip("-")
    token = token.replace(".", "").replace(",", ".")
    try:
        val = float(token)
        return -val if negative else val
    except ValueError:
        return None


def strip_dots(token):
    """Remove '.' separators from a plain integer-looking token, no decimal handling.
    Used for the Registr. number and the Documento numeric part, e.g. '8.269' -> '8269',
    '2.66' -> '266'."""
    if token is None:
        return ""
    return token.replace(".", "")


def looks_like_code_token(tok):
    """
    True if `tok` looks like an invoice/document CODE rather than ordinary description
    prose — used to decide whether a token preceding a Documento [date][docnum] pair
    belongs to Documento's word-prefix or is still part of Descrizione. Examples that
    should be True: '050512306T', '4/01-24350', 'FPR', '29/25'. Examples that should
    be False (ordinary Italian words): 'acquisto', 'Fattura', 'sottoscrizione'.
    """
    if not tok:
        return False
    if any(ch.isdigit() for ch in tok):
        return True
    if "/" in tok or "-" in tok:
        return True
    if tok.isupper() and len(tok) <= 6:
        return True
    return False


def is_header_row(cells):
    """Detect the column-header row so we know real data starts after it."""
    joined = " ".join((c or "") for c in cells).lower()
    joined = joined.replace(".", "").replace("°", "")
    hits = sum(1 for kw in HEADER_KEYWORDS if kw in joined)
    return hits >= 3


# --------------------------------------------------------------------------------------
# ZONE-BASED extraction: instead of guessing column boundaries from a text blob, use
# the HEADER row's actual word coordinates (x0/x1) to define each column's horizontal
# span on the page, then bin every data word into whichever column's span contains it.
# This is far more robust than regex-parsing a merged blob because DARE and AVERE end
# up in their own real zones — no more inferring the split from a SALDO delta.
# --------------------------------------------------------------------------------------

def _normalize_word(s):
    """Lowercase and strip punctuation, for loose header-word matching."""
    return re.sub(r'[^a-z0-9]', '', s.lower())


def cluster_words_into_lines(words, tolerance=3):
    """Group words into physical lines by rounding their 'top' (vertical) position.
    Returns a list of (top_key, [words sorted by x0]) sorted top-to-bottom."""
    lines = {}
    for w in words:
        key = round(w["top"] / tolerance) * tolerance
        lines.setdefault(key, []).append(w)
    return [(top, sorted(ws, key=lambda w: w["x0"])) for top, ws in sorted(lines.items())]


# Account code pattern printed in the title band just ABOVE the column header, e.g.
# "GE 1010103   CREDITI VS.SOCI PER VERSAMENTI DOVUTI" -> code is "GE 1010103":
# a short ALL-CAPS letter prefix followed by a numeric account id. Adjust here if
# other prefixes/formats appear in the ledger.
ACCOUNT_CODE_RE = re.compile(r'\b([A-Z]{1,5})\s+(\d{3,})\b')


def extract_page_code(page, header_top, debug=False):
    """
    Extract the account code (e.g. 'GE 1010103') for this page from the title line
    printed just ABOVE the column header row.

    Approach: take all words strictly above header_top, cluster them into lines,
    and scan those lines BOTTOM-UP (the code line sits immediately above the header,
    below any letterhead/company banner) for the ACCOUNT_CODE_RE pattern. Returns
    the matched code string, or "" if no code line is found on this page.
    """
    try:
        words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    except Exception:
        return ""
    if not words:
        return ""

    above = [w for w in words if w["top"] < header_top - 1]
    if not above:
        return ""

    lines = cluster_words_into_lines(above)
    for top, line_words in reversed(lines):  # bottom-up: nearest line above header first
        line_text = " ".join(w["text"] for w in line_words)
        m = ACCOUNT_CODE_RE.search(line_text)
        if m:
            code = f"{m.group(1)} {m.group(2)}"
            if debug:
                print(f"    Code line found at top={top}: '{line_text}' -> Code='{code}'")
            return code
        if debug:
            print(f"    Line above header at top={top} has no code pattern: '{line_text}'")

    return ""


def find_header_zones(page, debug=False):
    """
    Locate the header row on this page via word coordinates and compute contiguous
    x-coordinate zones for each target column (see COLUMN_HEADER_WORDS).

    Tries the header's own physical line first; if that doesn't yield a confident
    match (e.g. a column label like "Documento N° del e Prot." wrapped onto a second
    visual line in a narrow column), progressively merges in the next couple of
    lines' words and retries — sorted by x-position, so a wrapped label's words still
    land next to the rest of that column's header text.

    Returns (zones, header_top). zones is a list of (col_name, x_start, x_end)
    covering the full page width left-to-right, or None if the header row isn't
    found or its words can't be confidently matched even after trying to un-wrap it.
    If debug=True, also prints WHY each attempt failed (page-level diagnostics).
    """
    try:
        words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    except Exception as e:
        if debug:
            print(f"    extract_words() failed: {e}")
        return None, None
    if not words:
        return None, None

    clustered = cluster_words_into_lines(words)

    for idx, (top, line_words) in enumerate(clustered):
        cell_texts = [w["text"] for w in line_words]
        if not is_header_row(cell_texts):
            continue

        zones, reason = _match_header_words_to_zones(line_words)
        if zones:
            return zones, top
        if debug:
            print(f"    header candidate at top={top}: single-line match failed ({reason})")

        # Might be a WRAPPED header (a column label spilling onto the next line(s)).
        # Merge in subsequent lines' words (re-sorted by x-position) and retry.
        merged_top = top
        merged_words = list(line_words)
        for extra_top, extra_words in clustered[idx + 1: idx + 3]:
            merged_words = sorted(merged_words + extra_words, key=lambda w: w["x0"])
            merged_top = extra_top
            zones, reason = _match_header_words_to_zones(merged_words)
            if zones:
                return zones, merged_top
        if debug:
            print(f"    header candidate at top={top}: still failed after merging nearby lines ({reason})")

    return None, None


def _match_header_words_to_zones(line_words):
    """
    Given one header line's words (sorted left-to-right), match them against
    COLUMN_HEADER_WORDS in column order to find each column's core x-range, then
    expand into contiguous zones using midpoints between adjacent columns (so every
    data word, even if wider than the header label, still lands in exactly one zone).

    Matching is LENIENT: each expected word is searched for within the next few
    tokens (not required to be the immediate next token), tolerating a stray extra
    token or minor spacing quirk without failing the whole match.

    Returns (zones, failure_reason). zones is None if the expected words can't all
    be matched — failure_reason explains why, for diagnostics (falls back to the
    regex/blob approach for this page elsewhere).
    """
    pos = 0
    col_ranges = []
    failure_reason = None
    LOOKAHEAD = 3

    for col_name in DATA_FIELD_COLUMNS:
        expected = COLUMN_HEADER_WORDS.get(col_name, [])
        if not expected:
            continue

        matched = []
        cursor = pos
        ok = True
        for exp_word in expected:
            found_idx = None
            for j in range(cursor, min(cursor + LOOKAHEAD, len(line_words))):
                if _normalize_word(line_words[j]["text"]) == _normalize_word(exp_word):
                    found_idx = j
                    break
            if found_idx is None:
                ok = False
                failure_reason = (
                    f"column '{col_name}': expected header word '{exp_word}' not "
                    f"found within {LOOKAHEAD} tokens after position {cursor} "
                    f"(saw: {[w['text'] for w in line_words[cursor:cursor+LOOKAHEAD]]})"
                )
                break
            matched.append(line_words[found_idx])
            cursor = found_idx + 1

        if not ok or not matched:
            return None, failure_reason

        x0 = min(w["x0"] for w in matched)
        x1 = max(w["x1"] for w in matched)
        col_ranges.append((col_name, x0, x1))
        pos = cursor

    if len(col_ranges) != len(DATA_FIELD_COLUMNS):
        return None, failure_reason or "not all columns matched"

    zones = []
    for idx, (name, x0, x1) in enumerate(col_ranges):
        left = 0.0 if idx == 0 else (col_ranges[idx - 1][2] + x0) / 2
        right = 100000.0 if idx == len(col_ranges) - 1 else (x1 + col_ranges[idx + 1][1]) / 2
        zones.append((name, left, right))
    return zones, None


def bin_words_into_zone_rows(page, zones, header_top, page_num, page_code="", debug=False):
    """
    Take every word BELOW the header line and bin it into column zones per physical
    row, returning a list of raw dicts: {col_name: "joined text", ..., "_page": N} in
    top-to-bottom order. Each dict represents ONE physical line on the page
    (continuation lines from wrapped text are separate entries here —
    merge_continuation_rows() joins them, possibly across a page boundary).

    If debug=True, prints EVERY word's text, coordinates (x0, x1, center), and which
    zone it landed in (or "UNASSIGNED" if it fell outside all zones) — use this to
    see exactly whether a word's position matches the zone we computed for its column.
    """
    try:
        words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    except Exception:
        return []

    data_words = [w for w in words if w["top"] > header_top + 2]
    raw_rows = []

    if debug:
        print(f"    Zone boundaries for reference:")
        for name, left, right in zones:
            print(f"      {name}: [{left:.1f}, {right:.1f})")

    for top, line_words in cluster_words_into_lines(data_words):
        bucket = {name: [] for name, _, _ in zones}
        if debug:
            print(f"    -- line at top={top} --")
        for w in line_words:
            cx = (w["x0"] + w["x1"]) / 2
            assigned = None
            for name, left, right in zones:
                if left <= cx < right:
                    bucket[name].append(w["text"])
                    assigned = name
                    break
            if debug:
                tag = assigned if assigned else "** UNASSIGNED **"
                print(f"      word={w['text']!r:20} x0={w['x0']:.1f} x1={w['x1']:.1f} "
                      f"center={cx:.1f}  -> {tag}")
        row = {name: " ".join(vals).strip() for name, vals in bucket.items()}
        if any(v for v in row.values()):
            row["_page"] = page_num
            row["_code"] = page_code
            normalize_zone_overflow(row)
            raw_rows.append(row)

    return raw_rows


def normalize_zone_overflow(row):
    """
    Fix systematic zone-boundary overflow CONFIRMED via real coordinate debugging
    on the actual PDF (pages 9/69):

    1. The Registr. N° del zone captures the FIRST word of Descrizione, because the
       Descrizione DATA column physically starts (~x=162) to the LEFT of the
       Descrizione HEADER label (~x=185). Registr's real format is strictly
       [number-or-***][date] — so any tokens beyond that valid prefix are moved to
       the FRONT of Descrizione where they belong.
       e.g. Registr='6.233 30/04/2025 Fattura' -> Registr='6.233 30/04/2025',
            Descrizione='Fattura' + existing.
       Continuation lines whose lone stray word lands in Registr (e.g.
       'manutenzioni') get fully moved to Descrizione the same way.

    2. The Partita zone can capture the TAIL of Documento (e.g. a trailing '2025'
       of the protocol number printing at Partita's left edge). Partita's real
       format is slash-containing tokens — leading tokens WITHOUT a '/' are moved
       to the END of Documento.
       e.g. Partita='2025 482/2025' -> Documento += ' 2025', Partita='482/2025'.

    Mutates and returns `row` (raw string values, pre-merge).
    """
    # --- 1. Registr -> Descrizione overflow -------------------------------------
    reg_text = row.get("Registr. N° del", "").strip()
    if reg_text:
        tokens = reg_text.split()
        keep = 0
        if len(tokens) >= 1 and REGNUM_TOKEN_RE.match(tokens[0]):
            keep = 1
            if len(tokens) >= 2 and DATE_TOKEN_RE.match(tokens[1]):
                keep = 2
        overflow = tokens[keep:]
        if overflow:
            row["Registr. N° del"] = " ".join(tokens[:keep])
            existing = row.get("Descrizione Movimento", "").strip()
            row["Descrizione Movimento"] = " ".join(overflow + ([existing] if existing else []))

    # --- 2. Documento -> Partita overflow ---------------------------------------
    partita_text = row.get("Partita", "").strip()
    if partita_text:
        tokens = partita_text.split()
        moved = []
        while tokens and "/" not in tokens[0]:
            moved.append(tokens.pop(0))
        if moved:
            existing_doc = row.get("Documento N° del e Prot.", "").strip()
            row["Documento N° del e Prot."] = " ".join(
                ([existing_doc] if existing_doc else []) + moved
            )
            row["Partita"] = " ".join(tokens)

    return row


def strip_dots_in_plain_numbers(text):
    """Within a column's text, strip '.' from any standalone token that's a plain
    integer-with-dots (e.g. '8.269' -> '8269'), leaving dates, asterisks, and
    slash/dash codes untouched. Used for Registr. N° del and Documento N° del e Prot."""
    if not text:
        return text
    tokens = text.split()
    return " ".join(strip_dots(t) if DOCNUM_TOKEN_RE.match(t) else t for t in tokens)


def merge_continuation_rows(raw_rows, seed_row=None):
    """
    A wrapped/continuation line (e.g. a long Descrizione spilling onto a second
    physical line, or a SALDO value wrapping below) shows up as its own raw_row with
    an empty 'Data Comp.' zone. Merge such rows into the PREVIOUS real transaction
    row: append any non-empty zone text to that zone (with a space for text columns),
    or fill it in if it was previously blank (for numeric columns like SALDO).

    `seed_row`, if given, is the last (possibly still-incomplete) transaction row
    carried over from the PREVIOUS page — this is what lets a SALDO that wrapped
    onto the very first line of the next page still get merged correctly. The
    seed's own "_page" is preserved (it keeps the page where the transaction
    actually started), even though the continuation text came from a later page.

    Rows with no previous transaction to attach to (and no seed) are dropped (stray
    header/footer text). Returns the merged list of rows — the CALLER is responsible
    for holding back the last row if more continuation might still follow on the
    next page (see process_pdf()).
    """
    merged = [dict(seed_row)] if seed_row else []
    for row in raw_rows:
        has_date = bool(row.get("Data Comp.", "").strip())

        if not has_date:
            dare_txt = row.get("DARE", "").strip()
            avere_txt = row.get("AVERE", "").strip()
            saldo_txt = row.get("SALDO", "").strip()
            if dare_txt and avere_txt:
                # Page-end carry-forward/totals row (both DARE and AVERE populated,
                # no date) — this is NOT a continuation of the previous transaction;
                # drop it entirely per instructions (skip rows with both populated).
                continue

            # An orphan carrying ONLY amounts (no text content) can be either:
            #  (a) a wrapped-amounts continuation of an INCOMPLETE previous row
            #      (prev has no SALDO yet) -> merge it, or
            #  (b) a section/page totals line sitting under a COMPLETE previous row
            #      (prev already has its SALDO) -> skip it; merging would corrupt
            #      the previous row's amounts by concatenating totals onto them
            #      (confirmed real case: one-entry section on the last page whose
            #      totals duplicated the row's own 781.995,00 values).
            has_text_content = any(
                row.get(col, "").strip()
                for col in ("Registr. N° del", "Descrizione Movimento",
                            "Documento N° del e Prot.", "Partita")
            )
            amounts_only = (dare_txt or avere_txt or saldo_txt) and not has_text_content
            if amounts_only and merged and merged[-1].get("SALDO", "").strip():
                continue

        if has_date or not merged:
            merged.append(dict(row))
            continue

        prev = merged[-1]
        for col in DATA_FIELD_COLUMNS:
            val = row.get(col, "").strip()
            if not val:
                continue
            if prev.get(col, "").strip():
                prev[col] = f"{prev[col]} {val}".strip()
            else:
                prev[col] = val
    return merged


def _convert_amount_field(raw_text, field_name, issues):
    """
    Convert one amount field's raw zone text to a number.
    Returns (value, ok):
      - "" (empty text)        -> (None-equivalent blank handled by caller, ok)
      - single valid money txt -> (float, ok)
      - MULTIPLE money tokens  -> caused by an erroneous continuation merge; if all
        tokens are identical (same value read twice, common when a row repeats its
        amount on a wrapped line) take that value, else take the LAST token (most
        ledgers print the final/effective figure last) — either way, flag it.
      - unparseable text       -> (0, flagged) so the problem is VISIBLE in Status
        instead of silently becoming 0.
    """
    text = (raw_text or "").strip()
    if not text:
        return None, True

    val = parse_it_number(text)
    if val is not None:
        return val, True

    # Not a single clean number — check for multiple money tokens (merge artifact).
    tokens = MONEY_TOKEN_RE_FINDALL.findall(text)
    if tokens:
        vals = [parse_it_number(t) for t in tokens]
        vals = [v for v in vals if v is not None]
        if vals:
            if all(v == vals[0] for v in vals):
                issues.append(f"{field_name} had repeated value text '{text}', used {vals[0]}")
                return vals[0], False
            issues.append(f"{field_name} had MULTIPLE values '{text}', used last {vals[-1]} — verify")
            return vals[-1], False

    issues.append(f"{field_name} text '{text}' is not a parseable amount")
    return 0, False


def convert_zone_row(row, page_num):
    """Convert one raw zone-binned row (all string values) into the final typed row
    dict. Status is "Pass" only if every populated amount field parsed cleanly;
    otherwise Status carries a description of what went wrong (unparseable or
    multiple concatenated values), so nothing silently becomes 0.
    Uses row["_page"] as the Page value if present (preserves the page where a
    transaction actually started, even if merged continuation text came from a
    later page), otherwise falls back to the given page_num."""
    effective_page = row.get("_page", page_num)
    issues = []
    try:
        dare_val, _ = _convert_amount_field(row.get("DARE", ""), "DARE", issues)
        avere_val, _ = _convert_amount_field(row.get("AVERE", ""), "AVERE", issues)
        saldo_val, _ = _convert_amount_field(row.get("SALDO", ""), "SALDO", issues)

        dare = dare_val if dare_val is not None else 0
        avere = avere_val if avere_val is not None else 0
        saldo = saldo_val if saldo_val is not None else ""

        status = "Pass" if not issues else "Error: " + "; ".join(issues)

        return {
            "Page": effective_page,
            "Code": row.get("_code", ""),
            "Data Comp.": row.get("Data Comp.", "").strip(),
            "Registr. N° del": strip_dots_in_plain_numbers(row.get("Registr. N° del", "")),
            "Descrizione Movimento": row.get("Descrizione Movimento", "").strip(),
            "Documento N° del e Prot.": strip_dots_in_plain_numbers(row.get("Documento N° del e Prot.", "")),
            "Partita": row.get("Partita", "").strip(),
            "DARE": dare,
            "AVERE": avere,
            "SALDO": saldo,
            "Status": status,
        }
    except Exception as e:
        return {
            "Page": effective_page,
            "Code": row.get("_code", ""),
            "Data Comp.": row.get("Data Comp.", ""),
            "Registr. N° del": row.get("Registr. N° del", ""),
            "Descrizione Movimento": (row.get("Descrizione Movimento", "") or "")[:150],
            "Documento N° del e Prot.": row.get("Documento N° del e Prot.", ""),
            "Partita": row.get("Partita", ""),
            "DARE": 0,
            "AVERE": 0,
            "SALDO": "",
            "Status": f"Error: {e}",
        }


def assign_dare_avere(importo, saldo, running_saldo):
    """
    Infer whether `importo` belongs in DARE or AVERE by checking the direction the
    SALDO moved. ASSUMPTION (please confirm against your ledger convention):
        SALDO increased  -> DARE
        SALDO decreased  -> AVERE
    If your account type uses the opposite convention, swap the two branches below.
    """
    if saldo is None:
        # No SALDO to compare against; default to DARE and flag for manual review.
        return importo, 0, True  # (dare, avere, ambiguous_flag)

    delta = round(saldo - running_saldo, 2)
    if delta >= 0:
        return importo, 0, False
    else:
        return 0, importo, False


def parse_transaction_line(line):
    """
    Parse ONE physical transaction line into its fields using left-to-right /
    right-to-left token consumption (not one rigid regex), since Documento and
    Partita are OPTIONAL per row and a single monolithic pattern would fail to
    match whenever they're absent.

    Field rules (as specified):
      - Data: first date token, kept as-is (dd/mm/yy or dd/mm/yyyy).
      - Registr. N° del: a number (dots stripped, e.g. '8.269' -> '8269') or a
        '*****' placeholder, followed by a date -> stored as "<num> <date>".
      - Descrizione Movimento: free text in between, kept as one joined string
        (not further split/trimmed beyond single-spacing the tokens).
      - Documento N° del e Prot. (optional, 3 tokens): [alphanumeric word] [date]
        [plain number, dots stripped, e.g. '2.66' -> '266'] -> stored as one
        joined string; blank if this 3-token pattern isn't found.
      - Partita (optional, 1 token): alphanumeric + '/' + yyyy, e.g. 'AB123/2025';
        blank if not found.
      - DARE / AVERE / SALDO: trailing money-format tokens (Italian '.'=thousands,
        ','=decimal, e.g. '57.869,10' -> 57869.10), up to 3 trailing tokens.

    Returns a dict with raw extracted pieces (not yet DARE/AVERE-assigned):
        {"data": str, "registr": str, "descrizione": str, "documento": str,
         "partita": str, "amount_tokens": [float, ...]}
    or None if the line doesn't even start with a date (i.e. an orphan line).
    """
    tokens = line.split()
    if not tokens or not DATE_TOKEN_RE.match(tokens[0]):
        return None  # orphan line (no leading date)

    data_val = tokens[0]
    idx = 1

    # Registr. N° del: number/placeholder + date, if present right after Data.
    registr_val = ""
    if (idx + 1 < len(tokens)
            and REGNUM_TOKEN_RE.match(tokens[idx])
            and DATE_TOKEN_RE.match(tokens[idx + 1])):
        regnum_raw, regdate = tokens[idx], tokens[idx + 1]
        regnum_clean = regnum_raw if set(regnum_raw) <= {"*"} else strip_dots(regnum_raw)
        registr_val = f"{regnum_clean} {regdate}"
        idx += 2

    remainder_tokens = tokens[idx:]

    # From the END: trailing money tokens (DARE/AVERE/SALDO candidates), up to 3.
    amount_tokens = []
    end = len(remainder_tokens)
    while end > 0 and MONEY_TOKEN_RE.match(remainder_tokens[end - 1]) and len(amount_tokens) < 3:
        amount_tokens.insert(0, remainder_tokens[end - 1])
        end -= 1
    remainder_tokens = remainder_tokens[:end]
    amount_vals = [parse_it_number(a) for a in amount_tokens]

    # Partita: trailing token 'alnum/yyyy'-style, PLUS an optional preceding code-like
    # token (e.g. "FPR" before "17/25/2025"), mirroring Documento's word-prefix logic —
    # some ledgers prefix the Partita value with a short type code.
    partita_val = ""
    if remainder_tokens and PARTITA_TOKEN_RE.match(remainder_tokens[-1]):
        partita_tokens = [remainder_tokens[-1]]
        remainder_tokens = remainder_tokens[:-1]
        if (len(remainder_tokens) > 1  # always leave at least 1 token for Descrizione
                and looks_like_code_token(remainder_tokens[-1])
                and not (len(remainder_tokens) >= 2
                         and DATE_TOKEN_RE.match(remainder_tokens[-2])
                         and DOCNUM_TOKEN_RE.match(remainder_tokens[-1]))):
            # Only claim it if it's NOT actually the docnum of a Documento
            # [date][docnum] pair sitting right before Partita.
            partita_tokens.insert(0, remainder_tokens[-1])
            remainder_tokens = remainder_tokens[:-1]
        partita_val = " ".join(partita_tokens)

    # Documento N° del e Prot.: trailing pattern [word(s)][date][docnum]. The first
    # field ("word") can itself be MORE THAN ONE space-separated token (e.g. "FPR 29/25"
    # or "4/01-24350"). We walk BACKWARD from the [date][docnum] pair and only pull in
    # a preceding token if it looks_like_code_token() (contains a digit, '/', '-', or
    # is a short all-caps token) — ordinary Italian description words (e.g. "acquisto",
    # "Fattura") are NOT code-like and are correctly left behind in Descrizione. Capped
    # at DOCUMENTO_WORD_MAX_TOKENS and always leaves at least 1 token for Descrizione.
    documento_val = ""
    if (len(remainder_tokens) >= 3
            and DATE_TOKEN_RE.match(remainder_tokens[-2])
            and DOCNUM_TOKEN_RE.match(remainder_tokens[-1])):
        doc_date, doc_num_raw = remainder_tokens[-2], remainder_tokens[-1]
        before_pair = remainder_tokens[:-2]  # tokens before the [date][docnum] pair

        word_tokens = []
        idx = len(before_pair) - 1
        # Never consume the very last remaining token if that would leave nothing
        # at all for Descrizione (keep at least 1 token back).
        min_keep = 1
        while (idx >= 0
               and len(word_tokens) < DOCUMENTO_WORD_MAX_TOKENS
               and (len(before_pair) - len(word_tokens)) > min_keep
               and looks_like_code_token(before_pair[idx])):
            word_tokens.insert(0, before_pair[idx])
            idx -= 1

        if word_tokens:
            doc_word = " ".join(word_tokens)
            documento_val = f"{doc_word} {doc_date} {strip_dots(doc_num_raw)}"
            remainder_tokens = before_pair[:len(before_pair) - len(word_tokens)]
        else:
            # No code-like token immediately precedes the date+docnum pair; treat
            # Documento as not present (this [date][docnum] pair likely isn't real
            # Documento data, or Documento genuinely has no word prefix here).
            documento_val = ""

    descrizione_val = " ".join(remainder_tokens).strip()

    return {
        "data": data_val,
        "registr": registr_val,
        "descrizione": descrizione_val,
        "documento": documento_val,
        "partita": partita_val,
        "amount_tokens": amount_vals,
    }


def parse_transaction_lines(tagged_lines):
    """
    Parse a SEQUENCE of (page_num, line_text) tuples spanning the ENTIRE document
    (not just one page) into transaction rows. Processing the whole document as one
    continuous sequence — rather than page-by-page in isolation — means that when a
    transaction is the LAST one on a page and its SALDO wrapped onto the FIRST line
    of the NEXT page, the look-ahead recovery below still finds it (this was the gap
    that caused spurious "no SALDO available" warnings at page boundaries).

    Same behavior as before per line:
      - missing-SALDO recovery by borrowing from a following orphan line (now allowed
        to cross a page boundary)
      - orphan (no-date) lines with amount(s) but no date treated as the page-end
        totals row and skipped entirely
      - per-line error handling: exception -> row kept with Status="Error"; else "Pass"

    Returns (rows, warnings).
    """
    rows = []
    warnings = []
    running_saldo = 0.0
    i = 0
    n = len(tagged_lines)

    while i < n:
        page_num, line = tagged_lines[i]

        if line == HEADER_RESET_SENTINEL:
            # New account/client/supplier section detected (repeated header row) —
            # the running SALDO isn't comparable across sections, so reset it.
            running_saldo = 0.0
            i += 1
            continue

        status = "Pass"
        parsed = None
        dare, avere, saldo = 0, 0, ""

        try:
            parsed = parse_transaction_line(line)

            if parsed is None:
                # Orphan line (no leading date) -> page-end totals/carry-forward row.
                # Skipped silently per instructions.
                i += 1
                continue

            amounts = parsed["amount_tokens"]
            importo, saldo_val = None, None

            if len(amounts) >= 2:
                importo, saldo_val = amounts[0], amounts[-1]
            elif len(amounts) == 1:
                importo = amounts[0]
                # Try to recover a missing SALDO from the following orphan (no-date)
                # line — which may be on the SAME page or the FIRST line of the NEXT
                # page, since tagged_lines spans the whole document.
                if i + 1 < n:
                    next_page, next_line = tagged_lines[i + 1]
                    next_tokens = next_line.split()
                    next_is_orphan = not next_tokens or not DATE_TOKEN_RE.match(next_tokens[0])
                    if next_is_orphan:
                        next_amounts = MONEY_TOKEN_RE_FINDALL.findall(next_line)
                        if next_amounts:
                            saldo_val = parse_it_number(next_amounts[0])
                            cross_page_note = (
                                f" (recovered from page {next_page})" if next_page != page_num else ""
                            )
                            warnings.append(
                                f"[page {page_num}] Row '{parsed['data']} "
                                f"{parsed['descrizione'][:40]}': SALDO was missing, "
                                f"borrowed {saldo_val} from following line{cross_page_note} "
                                f"— please verify."
                            )
                            i += 1  # consume that orphan line

            dare, avere, ambiguous = assign_dare_avere(importo or 0, saldo_val, running_saldo)
            if ambiguous:
                warnings.append(
                    f"[page {page_num}] Row '{parsed['data']} {parsed['descrizione'][:40]}': "
                    f"no SALDO available to infer DARE vs AVERE — defaulted to DARE, "
                    f"please verify."
                )

            saldo = saldo_val if saldo_val is not None else ""
            if saldo_val is not None:
                running_saldo = saldo_val

            rows.append({
                "Page": page_num,
                "Data Comp.": parsed["data"],
                "Registr. N° del": parsed["registr"],
                "Descrizione Movimento": parsed["descrizione"],
                "Documento N° del e Prot.": parsed["documento"],
                "Partita": parsed["partita"],
                "DARE": dare,
                "AVERE": avere,
                "SALDO": saldo,
                "Status": status,
            })

        except Exception as e:
            warnings.append(f"[page {page_num}] ERROR parsing line '{line[:60]}': {e}")
            rows.append({
                "Page": page_num,
                "Data Comp.": parsed["data"] if parsed else "",
                "Registr. N° del": parsed["registr"] if parsed else "",
                "Descrizione Movimento": parsed["descrizione"] if parsed else line[:100],
                "Documento N° del e Prot.": parsed["documento"] if parsed else "",
                "Partita": parsed["partita"] if parsed else "",
                "DARE": 0,
                "AVERE": 0,
                "SALDO": "",
                "Status": "Error",
            })

        i += 1

    return rows, warnings


def is_page_total_row(row_dict):
    """
    Skip a row when:
      - BOTH DARE and AVERE are populated (non-zero) together, OR
      - DARE, AVERE, and SALDO are ALL populated together
    Both patterns indicate the page-end carry-forward / totals row rather than a
    genuine transaction (which only ever populates one of DARE/AVERE plus SALDO).
    (Kept as a safety net — the blob parser's orphan-line skip already filters most
    of these out at the line level.)
    """
    dare = row_dict.get("DARE")
    avere = row_dict.get("AVERE")
    saldo = row_dict.get("SALDO")

    dare_set = dare not in (None, 0, 0.0, "")
    avere_set = avere not in (None, 0, 0.0, "")
    saldo_set = saldo not in (None, 0, 0.0, "")

    if dare_set and avere_set:
        return True
    if dare_set and avere_set and saldo_set:
        return True
    return False


# --------------------------------------------------------------------------------------
# Table extraction + blob routing
# --------------------------------------------------------------------------------------
def collect_page_content(page, page_num, header_state):
    """
    Collect raw content from one page WITHOUT parsing transactions yet:
      - tagged_lines: list of (page_num, line_text) for blob content found on this
        page, to be parsed together with every other page's lines afterward (so
        cross-page SALDO recovery works — see parse_transaction_lines()).
      - direct_rows: rows that pdfplumber already split cleanly into columns (rare;
        these are complete already and don't need cross-page continuity).
    Returns (tagged_lines, direct_rows, page_warnings).
    """
    table_settings_options = [
        {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
        {"vertical_strategy": "text", "horizontal_strategy": "text"},
        {"vertical_strategy": "lines", "horizontal_strategy": "text"},
    ]

    tagged_lines = []
    direct_rows = []
    page_warnings = []

    for settings in table_settings_options:
        try:
            tables = page.extract_tables(settings)
        except Exception:
            continue
        if not tables:
            continue

        found_any = False
        for table in tables:
            for raw_row in table:
                if not raw_row or all((c is None or str(c).strip() == "") for c in raw_row):
                    continue
                cells = [(c or "").strip() for c in raw_row]

                if is_header_row(cells):
                    header_state["seen"] = True
                    # Mark this as a likely NEW ACCOUNT SECTION boundary — the running
                    # SALDO used to infer DARE vs AVERE should reset here, since a
                    # repeated header commonly means a new client/supplier sub-ledger
                    # is starting (their balances aren't related to the previous one).
                    tagged_lines.append((page_num, HEADER_RESET_SENTINEL))
                    continue

                if not header_state["seen"]:
                    continue  # pre-header content (titles, letterhead, etc.)

                non_empty_cells = [c for c in cells if c]
                if len(non_empty_cells) == 1:
                    new_lines = [l.strip() for l in non_empty_cells[0].split("\n") if l.strip()]
                    tagged_lines.extend((page_num, l) for l in new_lines)
                    found_any = True
                elif len(non_empty_cells) >= len(DATA_FIELD_COLUMNS) - 2:
                    # Row already split into ~columns; map directly — complete already.
                    try:
                        row = {col: "" for col in DATA_FIELD_COLUMNS}
                        for col, val in zip(DATA_FIELD_COLUMNS, non_empty_cells):
                            row[col] = val
                        row["DARE"] = parse_it_number(row.get("DARE", "")) or 0
                        row["AVERE"] = parse_it_number(row.get("AVERE", "")) or 0
                        row["SALDO"] = parse_it_number(row.get("SALDO", "")) or ""
                        row["Page"] = page_num
                        row["Status"] = "Pass"
                        direct_rows.append(row)
                    except Exception as e:
                        direct_rows.append({
                            "Page": page_num, "Data Comp.": "", "Registr. N° del": "",
                            "Descrizione Movimento": " ".join(non_empty_cells)[:100],
                            "Documento N° del e Prot.": "", "Partita": "",
                            "DARE": 0, "AVERE": 0, "SALDO": "", "Status": "Error",
                        })
                        page_warnings.append(f"[page {page_num}] ERROR mapping split row: {e}")
                    found_any = True
                else:
                    # Odd partial split — treat the joined text as one more line.
                    tagged_lines.append((page_num, " ".join(non_empty_cells)))
                    found_any = True

        if header_state["seen"] and found_any:
            break  # this settings variant worked; no need to try the others

    if not header_state["seen"]:
        page_warnings.append("column header row not yet found on this page.")

    return tagged_lines, direct_rows, page_warnings


def process_pdf(pdf_path, max_pages=None, debug_pages=None):
    debug_pages = set(debug_pages or [])
    zone_rows = []
    pending_zone_row = None  # last (possibly incomplete) row carried from previous page
    last_seen_code = ""      # carry-forward account code for pages without their own code line
    fallback_tagged_lines = []
    fallback_direct_rows = []
    header_state = {"seen": False}
    zone_pages, fallback_pages = 0, 0

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        pages_to_process = min(max_pages, total_pages) if max_pages else total_pages
        print(f"Opened '{pdf_path}' — {total_pages} page(s) detected. "
              f"Processing first {pages_to_process} page(s) only.")

        for i, page in enumerate(pdf.pages[:pages_to_process], start=1):
            if i % 50 == 0 or i == pages_to_process:
                print(f"Processing page {i}/{pages_to_process}...")

            is_debug = i in debug_pages
            if is_debug:
                print(f"\n--- DEBUG page {i} ---")
            zones, header_top = find_header_zones(page, debug=is_debug)
            if is_debug:
                print(f"  Zones found: {zones is not None}")
                if zones:
                    for name, left, right in zones:
                        print(f"    {name}: [{left:.1f}, {right:.1f})")

            if zones:
                zone_pages += 1
                header_state["seen"] = True
                page_code = extract_page_code(page, header_top, debug=is_debug)
                if page_code:
                    last_seen_code = page_code
                else:
                    # No code line on this page (e.g. continuation page of the same
                    # account) — inherit the most recent code seen.
                    page_code = last_seen_code
                if is_debug:
                    print(f"  Page code: {page_code!r}")
                raw_rows = bin_words_into_zone_rows(page, zones, header_top, i,
                                                    page_code=page_code, debug=is_debug)
                if is_debug:
                    print(f"  Raw zone-binned rows on this page ({len(raw_rows)}):")
                    for rr in raw_rows:
                        print(f"    {rr}")
                seed = pending_zone_row
                merged = merge_continuation_rows(raw_rows, seed_row=seed)
                if merged:
                    # Hold back the LAST row — it might still get continuation text
                    # from the NEXT page's first line — until we know otherwise.
                    complete, pending_zone_row = merged[:-1], merged[-1]
                    zone_rows.extend(convert_zone_row(r, i) for r in complete)
                else:
                    pending_zone_row = None
            else:
                if is_debug:
                    print("  No zones found on this page — will use regex/blob fallback.")
                # This page's header/zones couldn't be determined — finalize any
                # pending zone row (no cross-page continuation possible into a
                # fallback page) and process this page with the regex/blob method.
                if pending_zone_row:
                    zone_rows.append(convert_zone_row(pending_zone_row, pending_zone_row.get("_page", i)))
                    pending_zone_row = None
                fallback_pages += 1
                tagged_lines, direct_rows, page_warnings = collect_page_content(page, i, header_state)
                fallback_tagged_lines.extend(tagged_lines)
                fallback_direct_rows.extend(direct_rows)
                for w in page_warnings:
                    print(f"  [page {i}] WARNING: {w}")
                if is_debug:
                    print(f"  Fallback tagged {len(tagged_lines)} line(s) for regex parsing:")
                    for pnum, ln in tagged_lines:
                        print(f"    {ln}")

        if pending_zone_row:
            zone_rows.append(convert_zone_row(pending_zone_row, pending_zone_row.get("_page", pages_to_process)))

    print(f"\nZone-based extraction used on {zone_pages} page(s); "
          f"regex/blob fallback used on {fallback_pages} page(s).")

    if not header_state["seen"]:
        print("\nWARNING: the column header row was never detected anywhere in the "
              "document (neither via word coordinates nor extract_tables()). No rows "
              "were treated as data. Check HEADER_KEYWORDS / is_header_row() / "
              "COLUMN_HEADER_WORDS.")
        return []

    fallback_rows = []
    if fallback_tagged_lines:
        print(f"Parsing {len(fallback_tagged_lines)} fallback line(s) via regex/blob method...")
        fallback_rows, parse_warnings = parse_transaction_lines(fallback_tagged_lines)
        for w in parse_warnings:
            print(f"  WARNING: {w}")

    all_rows = zone_rows + fallback_rows + fallback_direct_rows
    all_rows.sort(key=lambda r: r.get("Page", 0))
    valid_rows = [r for r in all_rows if not is_page_total_row(r)]
    skipped = len(all_rows) - len(valid_rows)
    if skipped:
        print(f"Skipped {skipped} page-total/carry-forward row(s) across the document.")

    add_page_totals(valid_rows)
    print(f"Extracted {len(valid_rows)} transaction row(s) total.")
    return valid_rows


def add_page_totals(rows):
    """
    Compute per-page totals and stamp them onto EVERY row of that page:
      TOTAL_DARE  = sum of DARE for all rows on that page
      TOTAL_AVERE = sum of AVERE for all rows on that page
      TOTAL_SALDO = sum of SALDO for all rows on that page (blank SALDO counts as 0)
    e.g. if page 1 has 10 rows, all 10 rows carry the same three total values.
    Mutates `rows` in place.
    """
    def as_num(v):
        return v if isinstance(v, (int, float)) else 0

    totals = {}
    for r in rows:
        p = r.get("Page", 0)
        t = totals.setdefault(p, [0.0, 0.0, 0.0])
        t[0] += as_num(r.get("DARE"))
        t[1] += as_num(r.get("AVERE"))
        t[2] += as_num(r.get("SALDO"))

    for r in rows:
        t = totals.get(r.get("Page", 0), [0.0, 0.0, 0.0])
        r["TOTAL_DARE"] = round(t[0], 2)
        r["TOTAL_AVERE"] = round(t[1], 2)
        r["TOTAL_SALDO"] = round(t[2], 2)


# --------------------------------------------------------------------------------------
# Excel writing
# --------------------------------------------------------------------------------------
def write_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Giornale"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col_name in enumerate(COLUMNS, start=1):
            value = row.get(col_name, "")
            if col_name in ("DARE", "AVERE"):
                value = value if value not in (None, "") else 0
            # SALDO is left blank when not found (blank or negative are both valid).
            ws.cell(row=r_idx, column=c_idx, value=value)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        max_len = len(col_name)
        for row in rows:
            val = row.get(col_name, "")
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\nSaved {len(rows)} row(s) to '{output_path}'")


# --------------------------------------------------------------------------------------
# CLI entry point
# --------------------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Convert Italian ledger (Giornale) PDF to Excel.")
    parser.add_argument("input_pdf", help="Path to the input PDF file")
    parser.add_argument("output_xlsx", nargs="?", help="Path to the output Excel file (optional)")
    parser.add_argument("--pages", type=int, default=0,
                         help="Number of pages to process from the start of the PDF. "
                              "Default: 0 = process ALL pages.")
    parser.add_argument("--debug-pages", type=str, default="",
                         help="Comma-separated page numbers to print detailed zone-matching "
                              "diagnostics for, e.g. --debug-pages 9,69")
    args = parser.parse_args()

    input_path = Path(args.input_pdf)
    if not input_path.exists():
        print(f"ERROR: file not found: {input_path}")
        sys.exit(1)

    output_path = Path(args.output_xlsx) if args.output_xlsx else input_path.with_suffix(".xlsx")

    max_pages = args.pages if args.pages and args.pages > 0 else None
    debug_pages = [int(p.strip()) for p in args.debug_pages.split(",") if p.strip()]
    rows = process_pdf(str(input_path), max_pages=max_pages, debug_pages=debug_pages)

    if not rows:
        print("WARNING: No transaction rows were extracted. The PDF layout may need custom tuning.")
    write_excel(rows, str(output_path))


if __name__ == "__main__":
    main()
