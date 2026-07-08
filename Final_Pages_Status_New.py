#!/usr/bin/env python3
"""
giornale_to_excel.py
=====================
Converts an Italian accounting ledger ("Giornale") PDF into a structured Excel file.

REAL structure confirmed from extract_tables() output on the actual PDF:
  extract_tables() splits the HEADER row cleanly into 8 columns:
      ['Data Comp.', 'Registr. N. del', 'Descrizione Movimento',
       'Documento N. del e Prot.', 'Partita', 'DARE', 'AVERE', 'SALDO']
  ...but DATA rows collapse into a single blob in column 0, with every other
  column empty. Each physical line inside that blob is one transaction, e.g.:

      29/05/2025 ***** 29/05/2025 AUMENTO DI CAPITALE COME DA ASSEMBLEA 500.000,00 500.000,00
      09/09/2025 ***** 09/09/2025 sottoscrizione aumento di capitale [PERSON_1] 250.000,00 250.000,00
      09/09/2025 ***** 09/09/2025 sottoscrizione aumento di capitale [PERSON_2] 250.000,00
      500.000,00 500.000,00                      <- orphan line, no date: page-total row, SKIP

  So this script does NOT try to trust extract_tables()'s column splitting for data
  rows. Instead it takes the raw text blob and regex-parses each line itself.

Target columns (in order) — EXACTLY matching the real header row extracted from the
PDF (row[0] of extract_tables() output), not a custom split:
    Data Comp. | Registr. N° del | Descrizione Movimento | Documento N° del e Prot.
    | Partita | DARE | AVERE | SALDO

Design notes
------------
- 100% local / offline. Only pdfplumber (extraction) + openpyxl (writing). No network
  calls, safe for confidential client data.
- Extraction: pdfplumber's extract_tables() (lines/lines, text/text, lines/text tried
  in order) to get the header + locate the data blob; then per-line TOKEN-BASED
  parsing of the blob text itself (see parse_transaction_line()) — not one rigid
  regex, since Documento/Partita are optional per row.
- Field rules per transaction line:
    * Data:                  first date token, kept as-is.
    * Registr. N° del:       number (dots stripped, '8.269' -> '8269') or '*****'
                              placeholder + a date -> "<num> <date>".
    * Descrizione Movimento: free text in between, kept as one joined string.
    * Documento N° del e Prot. (optional): [word][date][number, dots stripped,
                              '2.66' -> '266'] -> one joined string, blank if absent.
    * Partita (optional):    single token 'alnum/yyyy', blank if absent.
    * DARE / AVERE / SALDO:  trailing Italian money tokens ('.'=thousands,
                              ','=decimal, '57.869,10' -> 57869.10).
- DARE vs AVERE assignment: since the single "Importo" figure has no column
  information left once it's inside one text blob, this script infers DARE vs AVERE
  by comparing the transaction's SALDO to the running SALDO carried from the previous
  transaction: if SALDO increases -> DARE, if it decreases -> AVERE.
  ASSUMPTION — please confirm this matches your ledger's convention; it's the most
  common one for Italian "mastrino" client/partner ledgers but can be inverted for
  some account types. See `assign_dare_avere()` — flip the two branches if wrong.
- Missing SALDO recovery: if a transaction line only has 1 trailing amount (SALDO
  missing, likely wrapped onto the next physical line), the script borrows the FIRST
  number of the following orphan line as that transaction's SALDO, then discards the
  rest of the orphan line. Flagged in the console output for manual spot-check.
- Skips a row when BOTH DARE and AVERE are populated together, or when DARE, AVERE,
  AND SALDO are all populated together — the page-end carry-forward/totals row, per
  your explicit instruction. See `is_page_total_row()`.
- Blank fields (Documento, Partita, missing amounts) are left blank/0 rather than
  guessed.
- Processes ALL pages of the PDF by default (use --pages N to limit to first N pages).

Usage
-----
    python giornale_to_excel.py input.pdf output.xlsx      (processes ALL pages)
    python giornale_to_excel.py input.pdf output.xlsx --pages N   (optional: limit to first N pages)

If output.xlsx is omitted, it defaults to the same name as input.pdf with .xlsx extension.
"""

import sys
import re
import argparse
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------------------
# Column definitions (matches the REAL header found via extract_tables())
# --------------------------------------------------------------------------------------
COLUMNS = [
    "Page",
    "Data Comp.",
    "Registr. N° del",
    "Descrizione Movimento",
    "Documento N° del e Prot.",
    "Partita",
    "DARE",
    "AVERE",
    "SALDO",
    "Status",
]

# --------------------------------------------------------------------------------------
# Regex helpers
# --------------------------------------------------------------------------------------
# Header detection keywords (loose match on the actual header wording)
HEADER_KEYWORDS = ["data", "comp", "registr", "descrizione", "movimento",
                   "documento", "prot", "partita", "dare", "avere", "saldo"]

# --- Token-level patterns for each field (see parse_transaction_line() for how
#     these are consumed left-to-right / right-to-left along a line) -------------
DATE_TOKEN_RE = re.compile(r'^\d{1,2}/\d{1,2}/\d{2,4}$')
REGNUM_TOKEN_RE = re.compile(r'^(?:\d[\d.]*|\*+)$')          # e.g. "8.269" or "*****"
DOCNUM_TOKEN_RE = re.compile(r'^\d[\d.]*$')                  # e.g. "2.66" -> "266" (dots stripped, no decimal comma)
PARTITA_TOKEN_RE = re.compile(r'^[A-Za-z0-9-]+(?:/[A-Za-z0-9-]+)+$')  # e.g. "AB123/2025", "26/2025/FE/2025", "4/01-24350/2025"
MONEY_TOKEN_RE = re.compile(r'^-?\d{1,3}(?:\.\d{3})*,\d{2}-?$')  # e.g. "57.869,10" -> 57869.10
MONEY_TOKEN_RE_FINDALL = re.compile(r'-?\d{1,3}(?:\.\d{3})*,\d{2}-?')  # non-anchored, for scanning full lines

# Cap on how many tokens Documento's first field ("word") can span, e.g. "FPR 29/25"
# is 2 tokens. ASSUMPTION based on observed samples — raise if real data needs more.
DOCUMENTO_WORD_MAX_TOKENS = 2

# Sentinel inserted into the tagged-line sequence at every detected header row, used
# to reset the running SALDO (see parse_transaction_lines()) since a repeated header
# likely marks the start of a new account/client/supplier section in the ledger.
HEADER_RESET_SENTINEL = "\x00__HEADER_RESET__\x00"


def parse_it_number(token):
    """Convert an Italian-formatted MONEY string ('1.234,56' or '1234,56-') to float.
    Used for DARE / AVERE / SALDO (dot = thousands separator, comma = decimal)."""
    if token is None:
        return None
    token = token.strip()
    if token == "" or token in {"-", "--"}:
        return None
    negative = token.endswith("-")
    token = token.rstrip("-")
    token = token.replace(".", "").replace(",", ".")
    try:
        val = float(token)
        return -val if negative else val
    except ValueError:
        return None


def strip_dots(token):
    """Remove '.' separators from a plain integer-looking token, no decimal handling.
    Used for the Registr. number and the Documento numeric part, e.g. '8.269' -> '8269',
    '2.66' -> '266'."""
    if token is None:
        return ""
    return token.replace(".", "")


def looks_like_code_token(tok):
    """
    True if `tok` looks like an invoice/document CODE rather than ordinary description
    prose — used to decide whether a token preceding a Documento [date][docnum] pair
    belongs to Documento's word-prefix or is still part of Descrizione. Examples that
    should be True: '050512306T', '4/01-24350', 'FPR', '29/25'. Examples that should
    be False (ordinary Italian words): 'acquisto', 'Fattura', 'sottoscrizione'.
    """
    if not tok:
        return False
    if any(ch.isdigit() for ch in tok):
        return True
    if "/" in tok or "-" in tok:
        return True
    if tok.isupper() and len(tok) <= 6:
        return True
    return False


def is_header_row(cells):
    """Detect the column-header row so we know real data starts after it."""
    joined = " ".join((c or "") for c in cells).lower()
    joined = joined.replace(".", "").replace("°", "")
    hits = sum(1 for kw in HEADER_KEYWORDS if kw in joined)
    return hits >= 3


def assign_dare_avere(importo, saldo, running_saldo):
    """
    Infer whether `importo` belongs in DARE or AVERE by checking the direction the
    SALDO moved. ASSUMPTION (please confirm against your ledger convention):
        SALDO increased  -> DARE
        SALDO decreased  -> AVERE
    If your account type uses the opposite convention, swap the two branches below.
    """
    if saldo is None:
        # No SALDO to compare against; default to DARE and flag for manual review.
        return importo, 0, True  # (dare, avere, ambiguous_flag)

    delta = round(saldo - running_saldo, 2)
    if delta >= 0:
        return importo, 0, False
    else:
        return 0, importo, False


def parse_transaction_line(line):
    """
    Parse ONE physical transaction line into its fields using left-to-right /
    right-to-left token consumption (not one rigid regex), since Documento and
    Partita are OPTIONAL per row and a single monolithic pattern would fail to
    match whenever they're absent.

    Field rules (as specified):
      - Data: first date token, kept as-is (dd/mm/yy or dd/mm/yyyy).
      - Registr. N° del: a number (dots stripped, e.g. '8.269' -> '8269') or a
        '*****' placeholder, followed by a date -> stored as "<num> <date>".
      - Descrizione Movimento: free text in between, kept as one joined string
        (not further split/trimmed beyond single-spacing the tokens).
      - Documento N° del e Prot. (optional, 3 tokens): [alphanumeric word] [date]
        [plain number, dots stripped, e.g. '2.66' -> '266'] -> stored as one
        joined string; blank if this 3-token pattern isn't found.
      - Partita (optional, 1 token): alphanumeric + '/' + yyyy, e.g. 'AB123/2025';
        blank if not found.
      - DARE / AVERE / SALDO: trailing money-format tokens (Italian '.'=thousands,
        ','=decimal, e.g. '57.869,10' -> 57869.10), up to 3 trailing tokens.

    Returns a dict with raw extracted pieces (not yet DARE/AVERE-assigned):
        {"data": str, "registr": str, "descrizione": str, "documento": str,
         "partita": str, "amount_tokens": [float, ...]}
    or None if the line doesn't even start with a date (i.e. an orphan line).
    """
    tokens = line.split()
    if not tokens or not DATE_TOKEN_RE.match(tokens[0]):
        return None  # orphan line (no leading date)

    data_val = tokens[0]
    idx = 1

    # Registr. N° del: number/placeholder + date, if present right after Data.
    registr_val = ""
    if (idx + 1 < len(tokens)
            and REGNUM_TOKEN_RE.match(tokens[idx])
            and DATE_TOKEN_RE.match(tokens[idx + 1])):
        regnum_raw, regdate = tokens[idx], tokens[idx + 1]
        regnum_clean = regnum_raw if set(regnum_raw) <= {"*"} else strip_dots(regnum_raw)
        registr_val = f"{regnum_clean} {regdate}"
        idx += 2

    remainder_tokens = tokens[idx:]

    # From the END: trailing money tokens (DARE/AVERE/SALDO candidates), up to 3.
    amount_tokens = []
    end = len(remainder_tokens)
    while end > 0 and MONEY_TOKEN_RE.match(remainder_tokens[end - 1]) and len(amount_tokens) < 3:
        amount_tokens.insert(0, remainder_tokens[end - 1])
        end -= 1
    remainder_tokens = remainder_tokens[:end]
    amount_vals = [parse_it_number(a) for a in amount_tokens]

    # Partita: single trailing token 'alnum/yyyy', if present.
    partita_val = ""
    if remainder_tokens and PARTITA_TOKEN_RE.match(remainder_tokens[-1]):
        partita_val = remainder_tokens[-1]
        remainder_tokens = remainder_tokens[:-1]

    # Documento N° del e Prot.: trailing pattern [word(s)][date][docnum]. The first
    # field ("word") can itself be MORE THAN ONE space-separated token (e.g. "FPR 29/25"
    # or "4/01-24350"). We walk BACKWARD from the [date][docnum] pair and only pull in
    # a preceding token if it looks_like_code_token() (contains a digit, '/', '-', or
    # is a short all-caps token) — ordinary Italian description words (e.g. "acquisto",
    # "Fattura") are NOT code-like and are correctly left behind in Descrizione. Capped
    # at DOCUMENTO_WORD_MAX_TOKENS and always leaves at least 1 token for Descrizione.
    documento_val = ""
    if (len(remainder_tokens) >= 3
            and DATE_TOKEN_RE.match(remainder_tokens[-2])
            and DOCNUM_TOKEN_RE.match(remainder_tokens[-1])):
        doc_date, doc_num_raw = remainder_tokens[-2], remainder_tokens[-1]
        before_pair = remainder_tokens[:-2]  # tokens before the [date][docnum] pair

        word_tokens = []
        idx = len(before_pair) - 1
        # Never consume the very last remaining token if that would leave nothing
        # at all for Descrizione (keep at least 1 token back).
        min_keep = 1
        while (idx >= 0
               and len(word_tokens) < DOCUMENTO_WORD_MAX_TOKENS
               and (len(before_pair) - len(word_tokens)) > min_keep
               and looks_like_code_token(before_pair[idx])):
            word_tokens.insert(0, before_pair[idx])
            idx -= 1

        if word_tokens:
            doc_word = " ".join(word_tokens)
            documento_val = f"{doc_word} {doc_date} {strip_dots(doc_num_raw)}"
            remainder_tokens = before_pair[:len(before_pair) - len(word_tokens)]
        else:
            # No code-like token immediately precedes the date+docnum pair; treat
            # Documento as not present (this [date][docnum] pair likely isn't real
            # Documento data, or Documento genuinely has no word prefix here).
            documento_val = ""

    descrizione_val = " ".join(remainder_tokens).strip()

    return {
        "data": data_val,
        "registr": registr_val,
        "descrizione": descrizione_val,
        "documento": documento_val,
        "partita": partita_val,
        "amount_tokens": amount_vals,
    }


def parse_transaction_lines(tagged_lines):
    """
    Parse a SEQUENCE of (page_num, line_text) tuples spanning the ENTIRE document
    (not just one page) into transaction rows. Processing the whole document as one
    continuous sequence — rather than page-by-page in isolation — means that when a
    transaction is the LAST one on a page and its SALDO wrapped onto the FIRST line
    of the NEXT page, the look-ahead recovery below still finds it (this was the gap
    that caused spurious "no SALDO available" warnings at page boundaries).

    Same behavior as before per line:
      - missing-SALDO recovery by borrowing from a following orphan line (now allowed
        to cross a page boundary)
      - orphan (no-date) lines with amount(s) but no date treated as the page-end
        totals row and skipped entirely
      - per-line error handling: exception -> row kept with Status="Error"; else "Pass"

    Returns (rows, warnings).
    """
    rows = []
    warnings = []
    running_saldo = 0.0
    i = 0
    n = len(tagged_lines)

    while i < n:
        page_num, line = tagged_lines[i]

        if line == HEADER_RESET_SENTINEL:
            # New account/client/supplier section detected (repeated header row) —
            # the running SALDO isn't comparable across sections, so reset it.
            running_saldo = 0.0
            i += 1
            continue

        status = "Pass"
        parsed = None
        dare, avere, saldo = 0, 0, ""

        try:
            parsed = parse_transaction_line(line)

            if parsed is None:
                # Orphan line (no leading date) -> page-end totals/carry-forward row.
                # Skipped silently per instructions.
                i += 1
                continue

            amounts = parsed["amount_tokens"]
            importo, saldo_val = None, None

            if len(amounts) >= 2:
                importo, saldo_val = amounts[0], amounts[-1]
            elif len(amounts) == 1:
                importo = amounts[0]
                # Try to recover a missing SALDO from the following orphan (no-date)
                # line — which may be on the SAME page or the FIRST line of the NEXT
                # page, since tagged_lines spans the whole document.
                if i + 1 < n:
                    next_page, next_line = tagged_lines[i + 1]
                    next_tokens = next_line.split()
                    next_is_orphan = not next_tokens or not DATE_TOKEN_RE.match(next_tokens[0])
                    if next_is_orphan:
                        next_amounts = MONEY_TOKEN_RE_FINDALL.findall(next_line)
                        if next_amounts:
                            saldo_val = parse_it_number(next_amounts[0])
                            cross_page_note = (
                                f" (recovered from page {next_page})" if next_page != page_num else ""
                            )
                            warnings.append(
                                f"[page {page_num}] Row '{parsed['data']} "
                                f"{parsed['descrizione'][:40]}': SALDO was missing, "
                                f"borrowed {saldo_val} from following line{cross_page_note} "
                                f"— please verify."
                            )
                            i += 1  # consume that orphan line

            dare, avere, ambiguous = assign_dare_avere(importo or 0, saldo_val, running_saldo)
            if ambiguous:
                warnings.append(
                    f"[page {page_num}] Row '{parsed['data']} {parsed['descrizione'][:40]}': "
                    f"no SALDO available to infer DARE vs AVERE — defaulted to DARE, "
                    f"please verify."
                )

            saldo = saldo_val if saldo_val is not None else ""
            if saldo_val is not None:
                running_saldo = saldo_val

            rows.append({
                "Page": page_num,
                "Data Comp.": parsed["data"],
                "Registr. N° del": parsed["registr"],
                "Descrizione Movimento": parsed["descrizione"],
                "Documento N° del e Prot.": parsed["documento"],
                "Partita": parsed["partita"],
                "DARE": dare,
                "AVERE": avere,
                "SALDO": saldo,
                "Status": status,
            })

        except Exception as e:
            warnings.append(f"[page {page_num}] ERROR parsing line '{line[:60]}': {e}")
            rows.append({
                "Page": page_num,
                "Data Comp.": parsed["data"] if parsed else "",
                "Registr. N° del": parsed["registr"] if parsed else "",
                "Descrizione Movimento": parsed["descrizione"] if parsed else line[:100],
                "Documento N° del e Prot.": parsed["documento"] if parsed else "",
                "Partita": parsed["partita"] if parsed else "",
                "DARE": 0,
                "AVERE": 0,
                "SALDO": "",
                "Status": "Error",
            })

        i += 1

    return rows, warnings


def is_page_total_row(row_dict):
    """
    Skip a row when:
      - BOTH DARE and AVERE are populated (non-zero) together, OR
      - DARE, AVERE, and SALDO are ALL populated together
    Both patterns indicate the page-end carry-forward / totals row rather than a
    genuine transaction (which only ever populates one of DARE/AVERE plus SALDO).
    (Kept as a safety net — the blob parser's orphan-line skip already filters most
    of these out at the line level.)
    """
    dare = row_dict.get("DARE")
    avere = row_dict.get("AVERE")
    saldo = row_dict.get("SALDO")

    dare_set = dare not in (None, 0, 0.0, "")
    avere_set = avere not in (None, 0, 0.0, "")
    saldo_set = saldo not in (None, 0, 0.0, "")

    if dare_set and avere_set:
        return True
    if dare_set and avere_set and saldo_set:
        return True
    return False


# --------------------------------------------------------------------------------------
# Table extraction + blob routing
# --------------------------------------------------------------------------------------
def collect_page_content(page, page_num, header_state):
    """
    Collect raw content from one page WITHOUT parsing transactions yet:
      - tagged_lines: list of (page_num, line_text) for blob content found on this
        page, to be parsed together with every other page's lines afterward (so
        cross-page SALDO recovery works — see parse_transaction_lines()).
      - direct_rows: rows that pdfplumber already split cleanly into columns (rare;
        these are complete already and don't need cross-page continuity).
    Returns (tagged_lines, direct_rows, page_warnings).
    """
    table_settings_options = [
        {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
        {"vertical_strategy": "text", "horizontal_strategy": "text"},
        {"vertical_strategy": "lines", "horizontal_strategy": "text"},
    ]

    tagged_lines = []
    direct_rows = []
    page_warnings = []
    DATA_FIELD_COLUMNS = [c for c in COLUMNS if c not in ("Page", "Status")]

    for settings in table_settings_options:
        try:
            tables = page.extract_tables(settings)
        except Exception:
            continue
        if not tables:
            continue

        found_any = False
        for table in tables:
            for raw_row in table:
                if not raw_row or all((c is None or str(c).strip() == "") for c in raw_row):
                    continue
                cells = [(c or "").strip() for c in raw_row]

                if is_header_row(cells):
                    header_state["seen"] = True
                    # Mark this as a likely NEW ACCOUNT SECTION boundary — the running
                    # SALDO used to infer DARE vs AVERE should reset here, since a
                    # repeated header commonly means a new client/supplier sub-ledger
                    # is starting (their balances aren't related to the previous one).
                    tagged_lines.append((page_num, HEADER_RESET_SENTINEL))
                    continue

                if not header_state["seen"]:
                    continue  # pre-header content (titles, letterhead, etc.)

                non_empty_cells = [c for c in cells if c]
                if len(non_empty_cells) == 1:
                    new_lines = [l.strip() for l in non_empty_cells[0].split("\n") if l.strip()]
                    tagged_lines.extend((page_num, l) for l in new_lines)
                    found_any = True
                elif len(non_empty_cells) >= len(DATA_FIELD_COLUMNS) - 2:
                    # Row already split into ~columns; map directly — complete already.
                    try:
                        row = {col: "" for col in DATA_FIELD_COLUMNS}
                        for col, val in zip(DATA_FIELD_COLUMNS, non_empty_cells):
                            row[col] = val
                        row["DARE"] = parse_it_number(row.get("DARE", "")) or 0
                        row["AVERE"] = parse_it_number(row.get("AVERE", "")) or 0
                        row["SALDO"] = parse_it_number(row.get("SALDO", "")) or ""
                        row["Page"] = page_num
                        row["Status"] = "Pass"
                        direct_rows.append(row)
                    except Exception as e:
                        direct_rows.append({
                            "Page": page_num, "Data Comp.": "", "Registr. N° del": "",
                            "Descrizione Movimento": " ".join(non_empty_cells)[:100],
                            "Documento N° del e Prot.": "", "Partita": "",
                            "DARE": 0, "AVERE": 0, "SALDO": "", "Status": "Error",
                        })
                        page_warnings.append(f"[page {page_num}] ERROR mapping split row: {e}")
                    found_any = True
                else:
                    # Odd partial split — treat the joined text as one more line.
                    tagged_lines.append((page_num, " ".join(non_empty_cells)))
                    found_any = True

        if header_state["seen"] and found_any:
            break  # this settings variant worked; no need to try the others

    if not header_state["seen"]:
        page_warnings.append("column header row not yet found on this page.")

    return tagged_lines, direct_rows, page_warnings


def process_pdf(pdf_path, max_pages=None):
    all_tagged_lines = []
    all_direct_rows = []
    header_state = {"seen": False}

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        pages_to_process = min(max_pages, total_pages) if max_pages else total_pages
        print(f"Opened '{pdf_path}' — {total_pages} page(s) detected. "
              f"Processing first {pages_to_process} page(s) only.")

        for i, page in enumerate(pdf.pages[:pages_to_process], start=1):
            if i % 50 == 0 or i == pages_to_process:
                print(f"Collecting page {i}/{pages_to_process}...")
            tagged_lines, direct_rows, page_warnings = collect_page_content(page, i, header_state)
            all_tagged_lines.extend(tagged_lines)
            all_direct_rows.extend(direct_rows)
            for w in page_warnings:
                print(f"  [page {i}] WARNING: {w}")

    if not header_state["seen"]:
        print("\nWARNING: the column header row was never detected anywhere in the "
              "document. No rows were treated as data. Check HEADER_KEYWORDS / "
              "is_header_row().")
        return []

    print(f"\nParsing {len(all_tagged_lines)} collected line(s) across the whole "
          f"document (enables cross-page SALDO recovery)...")
    parsed_rows, parse_warnings = parse_transaction_lines(all_tagged_lines)
    for w in parse_warnings:
        print(f"  WARNING: {w}")

    all_rows = parsed_rows + all_direct_rows
    all_rows.sort(key=lambda r: r.get("Page", 0))
    valid_rows = [r for r in all_rows if not is_page_total_row(r)]
    skipped = len(all_rows) - len(valid_rows)
    if skipped:
        print(f"Skipped {skipped} page-total/carry-forward row(s) across the document.")

    print(f"Extracted {len(valid_rows)} transaction row(s) total.")
    return valid_rows


# --------------------------------------------------------------------------------------
# Excel writing
# --------------------------------------------------------------------------------------
def write_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Giornale"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col_name in enumerate(COLUMNS, start=1):
            value = row.get(col_name, "")
            if col_name in ("DARE", "AVERE"):
                value = value if value not in (None, "") else 0
            # SALDO is left blank when not found (blank or negative are both valid).
            ws.cell(row=r_idx, column=c_idx, value=value)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        max_len = len(col_name)
        for row in rows:
            val = row.get(col_name, "")
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\nSaved {len(rows)} row(s) to '{output_path}'")


# --------------------------------------------------------------------------------------
# CLI entry point
# --------------------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Convert Italian ledger (Giornale) PDF to Excel.")
    parser.add_argument("input_pdf", help="Path to the input PDF file")
    parser.add_argument("output_xlsx", nargs="?", help="Path to the output Excel file (optional)")
    parser.add_argument("--pages", type=int, default=0,
                         help="Number of pages to process from the start of the PDF. "
                              "Default: 0 = process ALL pages.")
    args = parser.parse_args()

    input_path = Path(args.input_pdf)
    if not input_path.exists():
        print(f"ERROR: file not found: {input_path}")
        sys.exit(1)

    output_path = Path(args.output_xlsx) if args.output_xlsx else input_path.with_suffix(".xlsx")

    max_pages = args.pages if args.pages and args.pages > 0 else None
    rows = process_pdf(str(input_path), max_pages=max_pages)

    if not rows:
        print("WARNING: No transaction rows were extracted. The PDF layout may need custom tuning.")
    write_excel(rows, str(output_path))


if __name__ == "__main__":
    main()
